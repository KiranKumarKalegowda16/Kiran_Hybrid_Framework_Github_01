from openpyxl import load_workbook

def read_test_data(sheetname, testcase, file):

    """ Here read_test_data() function used to read the test data/inputs from excel workbook """

    data = []
    headers = None
    row_count = 0
    flag = True
    wb = load_workbook(file)  ## To load work book
    # print(list(wb)) --   iterator
    ws = wb[sheetname]  ##  We will get the specified sheet name in the file
    for row in ws.iter_rows(values_only=True):
        if flag:
            row_count += 1
        if row[0] == testcase and row[1].lower() == "yes":
            flag = False
            _temp = [_value for _value in row[2:] if _value]
            data.append(tuple(_temp))

    for row in ws.iter_rows(min_row=row_count - 1, values_only=True):
        _temp = [_value for _value in row[2:] if _value]
        headers = ",".join(_temp)
        break

    wb.close()
    if not headers or not data:
        raise ValueError("Empty headers or data")
    return [headers, data]


# header, data = read_test_data("new_sheet_one","nop_demo_login_page_data",r"C:\Users\kiran\OneDrive\Documents\new_test_excel_file.xlsx")
# print(header)
# print(data)